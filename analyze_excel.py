import openpyxl

excel_path = r'PfE-Stakeholder-and-Engagement-template (1).xlsx'
wb = openpyxl.load_workbook(excel_path)

# === STAKEHOLDER REGISTER ===
print("=" * 100)
print("📋 STAKEHOLDER REGISTER STRUCTURE")
print("=" * 100)

ws = wb['Stakeholder Register']

# Extract headers from row 2
headers = []
for col_num in range(1, 30):
    cell = ws.cell(row=2, column=col_num)
    if cell.value:
        headers.append(str(cell.value).strip())
    else:
        break

print(f"\n✅ COLUMNS ({len(headers)}):\n")
for i, h in enumerate(headers, 1):
    print(f"   {i:2d}. {h}")

# Sample data
print(f"\n📊 SAMPLE DATA (Row 3):\n")
if ws.cell(row=3, column=2).value:
    for i, header in enumerate(headers, 1):
        val = ws.cell(row=3, column=i).value
        if val:
            print(f"   • {header}: {str(val)[:60]}")

# === DROP DOWNS ===
print("\n" + "=" * 100)
print("📋 REFERENCE DATA - DROP DOWNS")
print("=" * 100)

wd = wb['Drop Downs']
dd_headers = []
for col_num in range(1, 10):
    cell = wd.cell(row=1, column=col_num)
    if cell.value:
        dd_headers.append(str(cell.value).strip())

print(f"\n✅ DROPDOWN CATEGORIES ({len(dd_headers)}):\n")
for i, h in enumerate(dd_headers, 1):
    print(f"   {i}. {h}")
    # Get unique values for this dropdown
    values = set()
    for row_num in range(2, 25):
        cell = wd.cell(row=row_num, column=i)
        if cell.value:
            val = str(cell.value).strip()
            if val and val not in values and len(values) < 8:
                values.add(val)
    if values:
        print(f"      Values: {', '.join(sorted(values)[:6])}")

# === ENGAGEMENT ACTIVITY PLAN ===
print("\n" + "=" * 100)
print("📋 ENGAGEMENT ACTIVITY PLAN")
print("=" * 100)

we = wb['Engagement activity plan']
print("\nStructure (first 12 rows):")
for row_num in range(1, 13):
    cell = we.cell(row=row_num, column=1)
    if cell.value:
        val = str(cell.value).strip()
        if len(val) > 80:
            print(f"   Row {row_num}: {val[:77]}...")
        else:
            print(f"   Row {row_num}: {val}")

print("\n" + "=" * 100)
print("🎯 DASHBOARD ENHANCEMENT OPPORTUNITIES")
print("=" * 100)
print("\n")
print("1️⃣  CONTACT MANAGEMENT")
print("   ✓ Add Contact Person field (currently missing)")
print("   ✓ Add Contact Details (Phone, Email, Website, Address)")
print("   ✓ Track primary & secondary contacts per stakeholder")
print()
print("2️⃣  PHASE-BASED ENGAGEMENT")
print("   ✓ Add project phase tracking (Phase 1, 2, 3)")
print("   ✓ Phase-specific notes per stakeholder")
print("   ✓ Timeline view: engagement across project lifecycle")
print()
print("3️⃣  RELATIONSHIP TRACKING")
print("   ✓ Add 'Relationship with project' field")
print("   ✓ Track: Involved, Affected, Interested")
print("   ✓ Show relationship type on dashboard")
print()
print("4️⃣  CONTRIBUTION OPPORTUNITIES")
print("   ✓ Add 'How stakeholder could contribute' field")
print("   ✓ Track capabilities: funding, expertise, advocacy, etc.")
print("   ✓ Filter: Who can help with what")
print()
print("5️⃣  ENGAGEMENT ACTIVITY TRACKING")
print("   ✓ Link to engagement activity plan")
print("   ✓ Track planned activities vs completed")
print("   ✓ Timeline: Next scheduled engagement activity")
print()
print("6️⃣  SENTIMENT & DETAILED NOTES")
print("   ✓ Separate 'Notes on Interest' & 'Notes on Influence'")
print("   ✓ Add detailed sentiment tracking per phase")
print("   ✓ Rich text notes with timestamps")
print()
print("7️⃣  STAKEHOLDER GROUPING")
print("   ✓ Add stakeholder group/category support")
print("   ✓ Group-level filtering & metrics")
print("   ✓ Group engagement status dashboard")
print()
print("8️⃣  GDPR & COMPLIANCE")
print("   ✓ Add data collection & consent tracking")
print("   ✓ Last contact date tracking")
print("   ✓ Data retention status indicators")
print()
